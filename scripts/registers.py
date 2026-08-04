"""Read the master register and push it into a working copy's mirror sheets.

README §6 makes ``registers/NEK_Master_Registers.xlsx`` the single source of
truth. The templates carry ``_EntityData`` / ``_BankAccounts`` sheets that are
labelled "interim mirror" of it, and nothing keeps the two in step.

Rather than detect that drift and fail, every run *overwrites* the working
copy's mirrors from the master register before filling anything. Master always
wins, so the two cannot disagree at the moment a document is produced.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ENTITY_SHEET = "01 Entity"
PROPERTY_SHEET = "02 Property & Lease"
BANK_SHEET = "09 Bank Accounts"
HEADER_ROW = 4  # data starts on the row below

MIRROR_ENTITY = "_EntityData"
MIRROR_BANK = "_BankAccounts"
MIRROR_CURRENCY = "_Currency"
MIRROR_FIRST_ROW = 3
# Lookup formulas span rows 3:100; anything past that is invisible to them.
MIRROR_LAST_ROW = 100
# The currency table is small and its formulas span 3:20.
CURRENCY_LAST_ROW = 20

# Values the register uses to mean "not applicable". Written through to the
# mirror they would print as a literal "Co. Reg. No. N/A" in the letterhead.
NULL_TOKENS = {"n/a", "na", "-", "--", "tbc", "[to confirm]"}

# Bank codes that stand for "not yet known". They load (see _key), but they must
# never reach a document: a reference like PV/WT/TBC/202609/001 is meaningless,
# and the account behind it is by definition unconfirmed.
PLACEHOLDER_BANK_CODES = {"TBC", "TBD", "TBA", "N/A", "NA", "-", "?", "XXX"}


class RegisterError(RuntimeError):
    """Raised when the register is missing, malformed, or over-full."""


@dataclass(frozen=True)
class Entity:
    code: str
    legal_name: str
    reg_no: str
    address: str
    currency: str
    tel: str
    email: str


@dataclass(frozen=True)
class BankAccount:
    entity_code: str
    bank_code: str
    bank_name: str
    currency: str
    account_name: str
    account_no: str

    @property
    def key(self) -> str:
        """The composite key the templates match on."""
        return f"{self.entity_code}|{self.bank_code}"

    @property
    def is_placeholder(self) -> bool:
        """True when the bank code is a stand-in rather than a real code."""
        return self.bank_code.upper() in PLACEHOLDER_BANK_CODES


def _key(value: object) -> str:
    """Normalise a *key* cell: trim only, never blank.

    Key fields must not go through :func:`_clean`. WT's bank code is literally
    ``TBC``, which is also a null token -- blanking it silently deleted the
    whole account from the register as the generation layer saw it. A row that
    quietly does not exist is far worse than a row that fails a later check.

    Whole numbers are rendered without a decimal point: Excel hands back unit
    "27" as the float ``27.0``, and ``"27.0"`` is not a key anyone typed.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean(value: object) -> str:
    """Normalise a *descriptive* cell, blanking placeholder tokens.

    Only for fields that are printed or ignored -- never for keys.
    """
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in NULL_TOKENS else text


def load_registers(register_path: str | Path) -> tuple[dict[str, Entity], dict[str, BankAccount]]:
    """Return ``({entity_code: Entity}, {key: BankAccount})`` from the master file."""
    path = Path(register_path).resolve()
    if not path.is_file():
        raise RegisterError(f"master register not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for required in (ENTITY_SHEET, BANK_SHEET):
            if required not in wb.sheetnames:
                raise RegisterError(f"{path.name} has no {required!r} sheet")

        entities: dict[str, Entity] = {}
        ws = wb[ENTITY_SHEET]
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            code = _key(row[0] if row else None)
            if not code:
                continue
            if code in entities:
                # Bank keys were already checked; entities were not, so a
                # duplicated code silently overwrote the earlier row and every
                # document for that entity would use whichever won.
                raise RegisterError(
                    f"duplicate entity code {code!r} in {ENTITY_SHEET}; "
                    "the entity code is the master key and must be unique"
                )
            entities[code] = Entity(
                code=code,
                legal_name=_clean(row[1]),
                reg_no=_clean(row[3]),      # D: Registration No.
                address=_clean(row[9]),     # J: Registered Address
                currency=_clean(row[6]),    # G: Functional Currency
                tel=_clean(row[14]),        # O: Tel
                email=_clean(row[15]),      # P: Email
            )

        accounts: dict[str, BankAccount] = {}
        ws = wb[BANK_SHEET]
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            bank_code = _key(row[0] if row else None)
            entity_code = _key(row[1]) if row and len(row) > 1 else ""
            if not bank_code or not entity_code:
                continue
            account = BankAccount(
                entity_code=entity_code,
                bank_code=bank_code,
                bank_name=_clean(row[2]),
                currency=_clean(row[6]),
                account_name=_clean(row[4]),
                account_no=_clean(row[5]),
            )
            if account.key in accounts:
                raise RegisterError(
                    f"duplicate bank key {account.key!r} in {BANK_SHEET}; "
                    "the (entity, bank) pair must be unique"
                )
            accounts[account.key] = account
    finally:
        wb.close()

    if not entities:
        raise RegisterError(f"no entities found in {ENTITY_SHEET}")
    if not accounts:
        raise RegisterError(f"no bank accounts found in {BANK_SHEET}")

    _assert_fits(len(entities), "entities")
    _assert_fits(len(accounts), "bank accounts")
    return entities, accounts


@dataclass(frozen=True)
class Property:
    code: str
    entity_code: str
    address: str
    unit: str
    tenant: str


def load_properties(register_path: str | Path) -> dict[str, Property]:
    """Return ``{normalised_key: Property}`` for the lettable units.

    Both the property code and the unit are indexed, so a caller can name
    either. Used to validate ``--unit`` on rental documents: filing by unit
    rather than by tenant keeps a private individual's name out of every
    filename and matches how rental income is actually reviewed.
    """
    path = Path(register_path).resolve()
    wb = load_workbook(path, data_only=True, read_only=True)
    index: dict[str, Property] = {}
    try:
        if PROPERTY_SHEET not in wb.sheetnames:
            return index
        for row in wb[PROPERTY_SHEET].iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            code = _key(row[0] if row else None)
            if not code:
                continue
            prop = Property(
                code=code,
                entity_code=_key(row[1]) if len(row) > 1 else "",
                address=_clean(row[2]) if len(row) > 2 else "",
                unit=_key(row[3]) if len(row) > 3 else "",
                tenant=_clean(row[5]) if len(row) > 5 else "",
            )
            for alias in (prop.code, prop.unit):
                if alias:
                    index.setdefault(_normalise(alias), prop)
    finally:
        wb.close()
    return index


def _normalise(text: str) -> str:
    """Letters and digits only, upper case -- for tolerant key matching."""
    return "".join(ch for ch in str(text).upper() if ch.isalnum())


def _assert_fits(count: int, what: str) -> None:
    capacity = MIRROR_LAST_ROW - MIRROR_FIRST_ROW + 1
    if count > capacity:
        raise RegisterError(
            f"{count} {what} exceed the {capacity}-row mirror capacity "
            f"(rows {MIRROR_FIRST_ROW}-{MIRROR_LAST_ROW}). Widen the lookup "
            "ranges in every document sheet before adding more."
        )


def sync_mirrors(wb, entities: dict[str, Entity], accounts: dict[str, BankAccount]) -> None:
    """Overwrite an open workbook's mirror sheets from the register.

    ``wb`` is a live Excel COM workbook -- the per-run working copy, never the
    template on disk.
    """
    ws = wb.Worksheets(MIRROR_ENTITY)
    _clear_mirror(ws, columns=7)
    for offset, entity in enumerate(sorted(entities.values(), key=lambda e: e.code)):
        row = MIRROR_FIRST_ROW + offset
        ws.Cells(row, 1).Value = entity.code
        ws.Cells(row, 2).Value = entity.legal_name
        ws.Cells(row, 3).Value = entity.reg_no
        ws.Cells(row, 4).Value = entity.address
        ws.Cells(row, 5).Value = entity.currency
        ws.Cells(row, 6).Value = entity.tel
        ws.Cells(row, 7).Value = entity.email

    ws = wb.Worksheets(MIRROR_BANK)
    _clear_mirror(ws, columns=7)
    for offset, account in enumerate(sorted(accounts.values(), key=lambda a: a.key)):
        row = MIRROR_FIRST_ROW + offset
        ws.Cells(row, 1).Value = account.key
        ws.Cells(row, 2).Value = account.bank_code
        ws.Cells(row, 3).Value = account.entity_code
        ws.Cells(row, 4).Value = account.bank_name
        ws.Cells(row, 5).Value = account.currency
        ws.Cells(row, 6).Value = account.account_name
        ws.Cells(row, 7).Value = account.account_no


def sync_currency(wb) -> None:
    """Write the currency wording table from :mod:`amount_in_words`.

    The template prints the major-unit name ("Ringgit Malaysia :") while the
    script writes the words that follow it. If those two disagreed, a document
    would read "Hong Kong Dollars : Two Thousand Ringgit ...". Rather than keep
    the wording in two places and hope, the Python module is the source and the
    sheet is rewritten from it on every run.
    """
    from amount_in_words import CURRENCIES  # local import: avoids a cycle

    ws = wb.Worksheets(MIRROR_CURRENCY)
    ws.Range(
        ws.Cells(MIRROR_FIRST_ROW, 1), ws.Cells(CURRENCY_LAST_ROW, 3)
    ).ClearContents()
    for offset, (code, (major, minor)) in enumerate(sorted(CURRENCIES.items())):
        row = MIRROR_FIRST_ROW + offset
        ws.Cells(row, 1).Value = code
        ws.Cells(row, 2).Value = major
        ws.Cells(row, 3).Value = minor


def _clear_mirror(ws, columns: int) -> None:
    """Blank the whole mirror band so removed rows cannot linger."""
    ws.Range(
        ws.Cells(MIRROR_FIRST_ROW, 1),
        ws.Cells(MIRROR_LAST_ROW, columns),
    ).ClearContents()
