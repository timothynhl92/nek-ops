"""Standing consistency audit of the master register.

Run this before any batch of documents, and after any hand-edit to the
register. It is deliberately broader than what the generation layer checks at
run time: the generator only validates the one voucher in front of it, while
this looks at the whole file and at problems that would not surface until some
future document happens to touch a bad row.

The failure mode it exists to prevent is the quiet one -- a row that never
matches, a code that silently sanitises to nothing, an entity referenced by a
property but absent from the entity register. Those produce plausible-looking
documents rather than errors.

    python scripts/audit_registers.py

Exit code 0 = no errors (warnings may still be printed), 1 = errors found.
"""

from __future__ import annotations

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from ref_and_filename import (  # noqa: E402
    DOC_TYPES,
    MAX_PATH,
    ReferenceError,
    counterparty_token,
    load_vendor_index,
)
from registers import (  # noqa: E402
    MIRROR_FIRST_ROW,
    MIRROR_LAST_ROW,
    PLACEHOLDER_BANK_CODES,
    load_registers,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
REGISTER = REPO_ROOT / "registers" / "NEK_Master_Registers.xlsx"
TEMPLATE = REPO_ROOT / "templates" / "NEK_Document_Templates.xlsx"

ENTITY_SHEET = "01 Entity"
PROPERTY_SHEET = "02 Property & Lease"
RECURRING_SHEET = "04 Recurring Payments"
VENDOR_SHEET = "05 Vendor"
CODE_SHEET = "08 Code Lists"
BANK_SHEET = "09 Bank Accounts"
HEADER_ROW = 4


class Report:
    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def error(self, where: str, message: str) -> None:
        self.errors.append(f"{where}: {message}")

    def warn(self, where: str, message: str) -> None:
        self.warnings.append(f"{where}: {message}")

    def section(self, title: str) -> None:
        print(f"\n{title}")
        print("-" * len(title))

    def ok(self, message: str) -> None:
        print(f"  ok    {message}")

    def note(self, message: str) -> None:
        print(f"        {message}")


def _column(ws, col: int, first_row: int = HEADER_ROW + 1) -> list[tuple[int, object]]:
    """Return ``(row, value)`` for a column, stopping at the first blank key."""
    out = []
    row = first_row
    blanks = 0
    while blanks < 3:
        value = ws.cell(row=row, column=col).value
        if value is None or str(value).strip() == "":
            blanks += 1
        else:
            blanks = 0
            out.append((row, value))
        row += 1
    return out


def check_key_hygiene(wb, report: Report) -> None:
    """Whitespace and casing in key fields.

    A trailing space is invisible in Excel and breaks every exact match that
    depends on the key. This register shipped with one.
    """
    report.section("Key hygiene")
    targets = [
        (ENTITY_SHEET, 1, "entity code"),
        (PROPERTY_SHEET, 1, "property code"),
        (PROPERTY_SHEET, 2, "owning entity"),
        (RECURRING_SHEET, 2, "entity"),
        (BANK_SHEET, 1, "bank code"),
        (BANK_SHEET, 2, "entity code"),
        (VENDOR_SHEET, 1, "vendor code"),
    ]
    dirty = 0
    for sheet, col, label in targets:
        for row, value in _column(wb[sheet], col):
            text = str(value)
            if text != text.strip():
                report.error(f"{sheet}!R{row}C{col}", f"{label} {text!r} has stray whitespace")
                dirty += 1
    if not dirty:
        report.ok("no stray whitespace in any key field")


def check_referential_integrity(wb, entities, accounts, report: Report) -> None:
    report.section("Referential integrity")
    known = set(entities)

    for sheet, col, label in (
        (PROPERTY_SHEET, 2, "owning entity"),
        (RECURRING_SHEET, 2, "entity"),
        (BANK_SHEET, 2, "entity code"),
    ):
        bad = [
            (row, value)
            for row, value in _column(wb[sheet], col)
            if str(value).strip() not in known
        ]
        if bad:
            for row, value in bad:
                report.error(f"{sheet}!row {row}", f"{label} {str(value)!r} is not in {ENTITY_SHEET}")
        else:
            report.ok(f"every {label} in {sheet} exists in {ENTITY_SHEET}")

    # Duplicates in anything used as a key.
    for sheet, col, label in (
        (ENTITY_SHEET, 1, "entity code"),
        (PROPERTY_SHEET, 1, "property code"),
        (VENDOR_SHEET, 1, "vendor code"),
    ):
        counts = Counter(str(v).strip() for _, v in _column(wb[sheet], col))
        dupes = {k: n for k, n in counts.items() if n > 1}
        if dupes:
            report.error(sheet, f"duplicate {label}s: {dupes}")
        else:
            report.ok(f"{label}s in {sheet} are unique")

    keys = Counter(a.key for a in accounts.values())
    dupes = {k: n for k, n in keys.items() if n > 1}
    if dupes:
        report.error(BANK_SHEET, f"duplicate (entity, bank) keys: {dupes}")
    else:
        report.ok("(entity, bank) keys are unique")


def check_filename_safety(report: Report) -> None:
    """Would every code survive becoming part of a filename?

    A vendor code that sanitises to something different -- or to nothing --
    produces filenames that cannot be traced back to the register.
    """
    report.section("Filename safety")
    index = load_vendor_index(REGISTER)
    codes = sorted(set(index.values()))
    unstable = []
    for code in codes:
        try:
            token = counterparty_token(code)
        except ReferenceError as exc:
            report.error(VENDOR_SHEET, f"vendor code {code!r} cannot form a token: {exc}")
            continue
        if token != code:
            unstable.append((code, token))
    if unstable:
        for code, token in unstable:
            report.error(VENDOR_SHEET, f"vendor code {code!r} sanitises to {token!r}")
    elif codes:
        report.ok(f"all {len(codes)} vendor codes round-trip unchanged")

    longest = max((len(c) for c in codes), default=0)
    sample = REPO_ROOT / "output" / (
        f"2026-09-01_SBOXCAP_PV_{'X' * longest}_PV-SBOXCAP-BOCOM-202609-001.pdf"
    )
    if len(str(sample)) > MAX_PATH:
        report.warn("filenames", f"worst-case path is {len(str(sample))} chars, over {MAX_PATH}")
    else:
        report.ok(f"worst-case path {len(str(sample))} chars, within {MAX_PATH}")


def check_bank_accounts(accounts, report: Report) -> None:
    report.section("Bank accounts")
    placeholders = [a.key for a in accounts.values() if a.is_placeholder]
    if placeholders:
        report.warn(
            BANK_SHEET,
            f"placeholder bank codes (documents are refused for these): {', '.join(sorted(placeholders))}",
        )
    else:
        report.ok("no placeholder bank codes")

    missing_currency = [a.key for a in accounts.values() if not a.currency]
    if missing_currency:
        report.error(BANK_SHEET, f"accounts with no currency: {missing_currency}")
    else:
        report.ok("every account has a currency")


def check_orphans(wb, entities, accounts, report: Report) -> None:
    """Rows nothing refers to. Not errors, but they rot quietly."""
    report.section("Orphans and unused rows")

    active: set[str] = set()
    for _, value in _column(wb[RECURRING_SHEET], 2):
        active.add(str(value).strip())
    for _, value in _column(wb[PROPERTY_SHEET], 2):
        active.add(str(value).strip())

    # An entity deliberately marked Dormant is expected to have no activity.
    # Warning about it every run would be noise, and noise is what stops people
    # reading audits at all.
    dormant = {
        str(wb[ENTITY_SHEET].cell(row=row, column=1).value).strip()
        for row, status in _column(wb[ENTITY_SHEET], 13)
        if str(status).strip().lower() != "active"
    }
    if dormant:
        report.note(f"non-active entities excluded from orphan checks: {', '.join(sorted(dormant))}")

    idle_entities = sorted(set(entities) - active - dormant)
    if idle_entities:
        report.warn(
            ENTITY_SHEET,
            f"active entities with no property and no recurring payment: {', '.join(idle_entities)}",
        )
    else:
        report.ok("every active entity has activity")

    by_entity = defaultdict(list)
    for account in accounts.values():
        by_entity[account.entity_code].append(account.bank_code)
    idle_accounts = sorted(k for k in by_entity if k not in active and k not in dormant)
    if idle_accounts:
        report.warn(
            BANK_SHEET,
            f"bank accounts for entities with no activity: {', '.join(idle_accounts)}",
        )
    else:
        report.ok("every bank account belongs to an active entity")


def check_doc_types(wb, report: Report) -> None:
    report.section("Document type codes")
    listed = {str(v).strip() for _, v in _column(wb[CODE_SHEET], 1)}
    missing = DOC_TYPES - listed
    if missing:
        report.error(CODE_SHEET, f"generated doc types absent from the code list: {sorted(missing)}")
    else:
        report.ok(f"all generated doc types present: {', '.join(sorted(DOC_TYPES))}")


def check_mirror_capacity(entities, accounts, report: Report) -> None:
    report.section("Template mirror capacity")
    capacity = MIRROR_LAST_ROW - MIRROR_FIRST_ROW + 1
    for label, count in (("entities", len(entities)), ("bank accounts", len(accounts))):
        if count > capacity:
            report.error("mirrors", f"{count} {label} exceed {capacity}-row capacity")
        else:
            report.ok(f"{count} {label} fit the {capacity}-row mirror band")

    wb = load_workbook(TEMPLATE)
    narrow = []
    for name in wb.sheetnames:
        if name.startswith("_") or name.startswith("00"):
            continue
        for row in wb[name].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and ("_EntityData" in v or "_BankAccounts" in v):
                    if f"${MIRROR_LAST_ROW}" not in v:
                        narrow.append(f"{name}!{cell.coordinate}")
    if narrow:
        report.error("template", f"lookup ranges not widened to row {MIRROR_LAST_ROW}: {narrow}")
    else:
        report.ok(f"every lookup range reaches row {MIRROR_LAST_ROW}")


def main() -> int:
    print(f"Auditing {REGISTER.name}")
    report = Report()
    wb = load_workbook(REGISTER, data_only=True)
    entities, accounts = load_registers(REGISTER)

    check_key_hygiene(wb, report)
    check_referential_integrity(wb, entities, accounts, report)
    check_bank_accounts(accounts, report)
    check_filename_safety(report)
    check_doc_types(wb, report)
    check_mirror_capacity(entities, accounts, report)
    check_orphans(wb, entities, accounts, report)

    print("\n" + "=" * 70)
    for warning in report.warnings:
        print(f"  WARN  {warning}")
    for error in report.errors:
        print(f"  ERROR {error}")
    print("=" * 70)
    print(
        f"{len(report.errors)} error(s), {len(report.warnings)} warning(s) "
        f"across {len(entities)} entities and {len(accounts)} accounts"
    )
    return 1 if report.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
