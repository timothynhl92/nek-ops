"""Read the recurring payment register and work out what falls due when.

The basis for the monthly closing checklist (README §6, §11 item 5).

The register records timing two ways: a day-of-month number, or free text like
"December 31" or "February 28 and August 31". Many rows record "N/A", which
means the timing was never captured — not that the item is unscheduled. Those
are reported separately rather than guessed at, because a checklist that
quietly omits an annual audit fee is worse than one that says it does not know
when it falls due.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

RECURRING_SHEET = "04 Recurring Payments"
HEADER_ROW = 4

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}
NULL_TOKENS = {"", "n/a", "na", "-", "tbc", "none"}

# Frequencies that recur every month by definition.
MONTHLY = {"monthly"}
# Frequencies that recur, but need an anchor month we do not hold.
PERIODIC = {"bi-monthly", "quarterly", "semi-annual", "annual"}


@dataclass(frozen=True)
class Recurring:
    ref: str
    entity: str
    direction: str          # Payable / Receivable
    category: str
    counterparty: str
    description: str
    currency: str
    amount_text: str        # as recorded -- not always a number
    amount: Decimal | None  # parsed, or None when the register holds prose
    frequency: str
    due_text: str
    due_day: int | None
    due_months: frozenset[int]   # empty when unknown
    # Per-month day, where the register gives one. Quarterly entries are often
    # written "31 January, 30 April, 31 July, 30 October" -- month-end dates
    # that differ by month. Taking a single day for all four would report some
    # of them a day out.
    due_day_by_month: dict[int, int]
    method: str
    bank_account: str
    approver: str
    preparer: str
    supporting_doc: str
    status: str
    notes: str

    @property
    def timing_known(self) -> bool:
        """Can this item's month be determined at all?"""
        if self.frequency.lower() in MONTHLY:
            return self.due_day is not None
        return bool(self.due_months)

    @property
    def is_zero_value(self) -> bool:
        """A recorded obligation that costs nothing.

        HHIL's secretarial fee is RM 0 — as a foreign company it has no fee to
        pay. The row is worth keeping so the obligation stays visible, but it
        needs no due date, so it should not be nagged about as missing one.
        """
        return self.amount is not None and self.amount == 0

    def due_in(self, year: int, month: int) -> bool:
        """Does this item fall due in the given month?"""
        if not self.timing_known:
            return False
        if self.frequency.lower() in MONTHLY:
            return True
        return month in self.due_months

    def due_date(self, year: int, month: int) -> date | None:
        """The due date within the month, clamped to the month's length."""
        day = self.due_day_by_month.get(month, self.due_day)
        if day is None:
            return None
        import calendar

        last = calendar.monthrange(year, month)[1]
        return date(year, month, min(day, last))


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _amount(value: object) -> tuple[str, Decimal | None]:
    text = _text(value)
    if isinstance(value, (int, float)):
        return text, Decimal(str(value))
    try:
        return text, Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError):
        # e.g. "Around 400" -- keep the words, refuse to invent a number.
        return text, None


def parse_due(
    value: object, frequency: str
) -> tuple[int | None, frozenset[int], dict[int, int]]:
    """Return ``(day_of_month, months, day_by_month)`` from the Due Day cell.

    Handles the shapes the register actually uses: a bare number, "Every 15 of
    the month", "December 31", "31 January, 30 April, 31 July, 30 October".
    Where each month carries its own day, that pairing is preserved -- those
    quarterly entries are month-end dates and differ by month.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        day = int(value)
        if 1 <= day <= 31:
            # A bare day with a non-monthly frequency tells us the day but not
            # which months, so the months stay unknown.
            months = frozenset(range(1, 13)) if frequency.lower() in MONTHLY else frozenset()
            return day, months, {}
        return None, frozenset(), {}

    text = _text(value)
    if text.lower() in NULL_TOKENS:
        return None, frozenset(), {}

    # Try to read "<day> <month>" / "<month> <day>" pairs chunk by chunk.
    day_by_month: dict[int, int] = {}
    for chunk in re.split(r",|\band\b", text, flags=re.I):
        lowered = chunk.lower()
        found = [n for name, n in MONTH_NAMES.items() if name in lowered]
        days = [int(d) for d in re.findall(r"\b(\d{1,2})\b", chunk) if 1 <= int(d) <= 31]
        if len(found) == 1 and days:
            day_by_month[found[0]] = days[0]

    lowered = text.lower()
    months = frozenset(n for name, n in MONTH_NAMES.items() if name in lowered)
    days = [int(d) for d in re.findall(r"\b(\d{1,2})\b", text) if 1 <= int(d) <= 31]
    default_day = days[0] if days else None
    if day_by_month:
        months = frozenset(day_by_month)
        default_day = day_by_month[min(day_by_month)]
    return default_day, months, day_by_month


def load_recurring(register_path: str | Path) -> list[Recurring]:
    """Every active row of the recurring payment register."""
    path = Path(register_path).resolve()
    wb = load_workbook(path, data_only=True, read_only=True)
    items: list[Recurring] = []
    try:
        if RECURRING_SHEET not in wb.sheetnames:
            return items
        for row in wb[RECURRING_SHEET].iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            ref = _text(row[0] if row else None)
            if not ref:
                continue
            status = _text(row[18]) if len(row) > 18 else ""
            if status and status.lower() != "active":
                continue
            frequency = _text(row[9])
            amount_text, amount = _amount(row[7])
            day, months, day_by_month = parse_due(row[10], frequency)
            items.append(
                Recurring(
                    ref=ref,
                    entity=_text(row[1]),
                    direction=_text(row[2]),
                    category=_text(row[3]),
                    counterparty=_text(row[4]),
                    description=_text(row[5]),
                    currency=_text(row[6]),
                    amount_text=amount_text,
                    amount=amount,
                    frequency=frequency,
                    due_text=_text(row[10]),
                    due_day=day,
                    due_months=months,
                    due_day_by_month=day_by_month,
                    method=_text(row[11]),
                    bank_account=_text(row[12]),
                    approver=_text(row[13]),
                    preparer=_text(row[14]),
                    supporting_doc=_text(row[15]) if len(row) > 15 else "",
                    status=status,
                    notes=_text(row[19]) if len(row) > 19 else "",
                )
            )
    finally:
        wb.close()
    return items


def split_for_month(
    items: list[Recurring], year: int, month: int, entity: str | None = None
) -> tuple[list[Recurring], list[Recurring]]:
    """Return ``(due_this_month, timing_unknown)``.

    Anything whose timing cannot be determined is returned separately rather
    than dropped. It still has to be paid; the register just does not say when.
    """
    if entity:
        items = [i for i in items if i.entity.upper() == entity.upper()]
    due = [i for i in items if i.due_in(year, month)]
    # A zero-value obligation needs no due date, so an absent one is not a gap.
    unknown = [i for i in items if not i.timing_known and not i.is_zero_value]
    due.sort(key=lambda i: (i.due_day or 99, i.entity, i.counterparty))
    unknown.sort(key=lambda i: (i.frequency, i.category, i.counterparty, i.entity))
    return due, unknown
