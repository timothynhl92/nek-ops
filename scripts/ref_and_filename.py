"""Build the on-document reference and the on-disk filename (README §5).

The reference is what gets printed; the filename is what gets stored. They are
not the same string, and the difference is the point: a reference contains "/",
which cannot appear in a filename.

Counterparty token rule (fixed, so filenames stay predictable):

1. If the payee matches a **vendor code** or a **vendor name** in the Vendor
   register (case-insensitive, punctuation-insensitive), use that vendor code.
2. Otherwise derive a token from the payee name:
   a. Unicode is folded to ASCII ("Sdn Bhd" survives, accents are stripped).
   b. "&" becomes "AND" -- §5 forbids "&" in filenames but dropping it silently
      would turn "Ng & Sons" into "NG SONS".
   c. Uppercased; every remaining character outside A-Z 0-9 becomes a space.
   d. Runs of whitespace collapse to a single hyphen.
   e. Truncated to 40 characters on a hyphen boundary, so the total path stays
      inside the 200-character limit §5 sets.
3. An empty result is an error, never a silent blank field.

Rows in the Vendor register flagged as example data are ignored -- the shipped
register contains one such row ("EXAMPLE ROW - delete before use"), and matching
against it would stamp a placeholder code onto a real document.
"""

from __future__ import annotations

import re
import unicodedata
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

VENDOR_SHEET = "05 Vendor"
VENDOR_HEADER_ROW = 4
EXAMPLE_MARKER = "example row"

MAX_COUNTERPARTY = 40
MAX_PATH = 200

DOC_TYPES = {"PV", "RV", "OR", "INV", "SAL"}


class ReferenceError(ValueError):
    """Raised when a reference or filename cannot be built safely."""


def build_reference(doctype: str, entity: str, bank: str, doc_date: date, sequence: int) -> str:
    """``PV/NEK/BOC/202607/014`` -- the string printed on the document."""
    doctype = doctype.upper()
    if doctype not in DOC_TYPES:
        raise ReferenceError(f"unknown doc type {doctype!r}; known: {', '.join(sorted(DOC_TYPES))}")
    if not 1 <= sequence <= 999:
        # TEXT(K5,"000") silently widens past 999, which would break the
        # fixed-width reference format every downstream filter assumes.
        raise ReferenceError(f"sequence {sequence} outside 001-999")
    return f"{doctype}/{entity.upper()}/{bank.upper()}/{doc_date:%Y%m}/{sequence:03d}"


def reference_to_filename_token(reference: str) -> str:
    """``PV/NEK/BOC/202607/014`` -> ``PV-NEK-BOC-202607-014``."""
    return reference.replace("/", "-")


def _fold(text: str) -> str:
    """ASCII-fold and uppercase, mapping '&' to 'AND'."""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.replace("&", " AND ")
    return text.upper()


def _normalise_for_match(text: str) -> str:
    """Reduce a name to letters and digits only, for tolerant comparison."""
    return re.sub(r"[^A-Z0-9]", "", _fold(text))


def load_vendor_index(register_path: str | Path) -> dict[str, str]:
    """Map normalised vendor code *and* name to the vendor code."""
    path = Path(register_path).resolve()
    index: dict[str, str] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if VENDOR_SHEET not in wb.sheetnames:
            return index
        for row in wb[VENDOR_SHEET].iter_rows(min_row=VENDOR_HEADER_ROW + 1, values_only=True):
            if not row:
                continue
            code = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            notes = str(row[19] or "").strip() if len(row) > 19 else ""
            if not code or EXAMPLE_MARKER in notes.lower():
                continue
            index[_normalise_for_match(code)] = code
            if name:
                index[_normalise_for_match(name)] = code
    finally:
        wb.close()
    return index


def counterparty_token(payee: str, vendor_index: dict[str, str] | None = None) -> str:
    """Apply the rule documented at the top of this module."""
    if not payee or not payee.strip():
        raise ReferenceError("payee is empty; cannot build a counterparty token")

    if vendor_index:
        match = vendor_index.get(_normalise_for_match(payee))
        if match:
            return _sanitise(match)

    token = _sanitise(payee)
    if not token:
        raise ReferenceError(
            f"payee {payee!r} contains no usable characters for a filename"
        )
    return token


def _sanitise(text: str) -> str:
    folded = _fold(text)
    spaced = re.sub(r"[^A-Z0-9]+", " ", folded).strip()
    token = "-".join(spaced.split())
    if len(token) > MAX_COUNTERPARTY:
        token = token[:MAX_COUNTERPARTY]
        # Prefer cutting at a word boundary over mid-word.
        if "-" in token:
            token = token.rsplit("-", 1)[0]
    return token.strip("-")


def build_filename(
    doc_date: date,
    entity: str,
    doctype: str,
    counterparty: str,
    reference: str,
    extension: str = "pdf",
    prefix: str = "",
) -> str:
    """``2026-09-01_NEK_PV_KWSP_PV-NEK-BOC-202609-001.pdf`` (§5)."""
    name = "_".join(
        (
            f"{doc_date:%Y-%m-%d}",
            entity.upper(),
            doctype.upper(),
            counterparty,
            reference_to_filename_token(reference),
        )
    )
    filename = f"{prefix}{name}.{extension.lstrip('.')}"
    if any(ch in filename for ch in '/\\&%#$(),\'"'):
        raise ReferenceError(f"filename retains a forbidden character: {filename!r}")
    return filename


def check_path_length(path: str | Path) -> None:
    """§5 caps the total path at 200 characters."""
    resolved = str(Path(path).resolve())
    if len(resolved) > MAX_PATH:
        raise ReferenceError(
            f"path is {len(resolved)} characters, over the {MAX_PATH} limit:\n{resolved}"
        )
